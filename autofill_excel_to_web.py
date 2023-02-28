from selenium import webdriver
from openpyxl import load_workbook

# Copyright
print("Made by SeenKid : https://yannberlemont.ch/")

# Charger le fichier Excel
wb = load_workbook(filename='data.xlsx', read_only=True)

# Nom de la feuille excel
ws = wb['Sheet1']

# Ouvrir le navigateur
driver = webdriver.Chrome()

# Charger la page Web à remplir
driver.get('URL_SITE')

# Boucle à travers chaque ligne du fichier Excel
for row in ws.iter_rows(min_row=2, values_only=True):
    # Extraire les données de chaque colonne
    data1 = row[0]
    data2 = row[1]
    data3 = row[2]
    data4 = row[3]
    data5 = row[4]
    data6 = row[5]
    data7 = row[6]
    data8 = row[10]
    data9 = row[11]
    
    # Trouver les champs de saisie sur la page Web
    input1 = driver.find_element_by_name('')
    input2 = driver.find_element_by_name('')
    input3 = driver.find_element_by_name('')
    input4 = driver.find_element_by_name('')
    input5 = driver.find_element_by_name('')
    input6 = driver.find_element_by_name('')
    input7 = driver.find_element_by_name('')

    # Effacer les champs de saisie (au cas où ils contiennent des données préexistantes)
    input1.clear()
    input2.clear()
    input3.clear()
    input4.clear()
    input5.clear()
    input6.clear()
    input7.clear()
    
    # Remplir les champs de saisie avec les données de l'extraction Excel
    input1.send_keys(data1)
    input2.send_keys(data2)
    input3.send_keys(data3)
    input4.send_keys(data4)
    input5.send_keys(data5)
    input6.send_keys(data6)
    input7.send_keys(data7)
    
    # Soumettre le formulaire (si nécessaire)
    driver.find_element_by_name('submit').click()
    
print("L'opération s'est terminée.")
